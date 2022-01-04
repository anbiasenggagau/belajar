# buatlah nilai-nilai pada kolom 3 di file transactions.xlsx menjadi 90%-nya

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1'] # mengakses sheet pada file, pastikan untuk mengetik nama sheet yang benar
cell = sheet['a1'] # cara pertama untuk memanggil cell
cell = sheet.cell(1,1) # cara kedua untuk memanggil cell
print(cell.value)
print()
print()

maxrow = sheet.max_row # mendapatkan jumlah baris yang berisi value

for item in range(2, maxrow + 1):
    cell = sheet.cell(item,3) #mengakses seluruh value di kolom 3
    fix_price = cell.value * 0.9
    fix_cell = sheet.cell(item, 4) # mengakses seluruh value di kolom 4
    fix_cell.value = fix_price # memperbarui seluruh value di kolom 4

# membuat grafik
values_chart = Reference(sheet, #membuat refrensi, angka mana saja yang mau dijadikan tabel
                  min_row=2,
                  max_row=sheet.max_row,
                  min_col=4,
                  max_col=4)

chart = BarChart() # membuat grafik batang yang masih kosong
chart.add_data(values_chart)  # memasukkan angka untuk grafik
sheet.add_chart(chart, 'e2') # meletakkan titik posisi kanan atas dari grafik
wb.save('transactions2.xlsx') #menyimpan hasil program ke file baru, atau juga bisa menimpah file terdahulu